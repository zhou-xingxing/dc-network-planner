from __future__ import annotations

import io
from collections.abc import Sequence
from typing import Any, cast
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

TEMPLATE_HEADERS = [
    "区域名称\n(必填)",
    "网络平面类型\n(必填)",
    "作用域\n(选填，空值默认 Global)",
    "IP地址段(CIDR)\n(必填，如 10.0.0.0/24)",
    "VLAN ID\n(选填，1-4094)",
    "网关位置\n(选填)",
    "网关IP\n(选填)",
]

COLUMN_WIDTHS = [20, 18, 26, 30, 18, 30, 18]
EXPORT_HEADERS = [
    "区域",
    "网络平面类型",
    "父级网络平面类型",
    "作用域",
    "是否私网",
    "VRF",
    "IP地址段(CIDR)",
    "VLAN ID",
    "网关位置",
    "网关IP",
    "更新时间",
]
EXPORT_COLUMN_WIDTHS = [20, 18, 18, 16, 10, 16, 20, 12, 30, 18, 28]
TEMPLATE_OPTION_SHEET = "候选项"
IMPORT_MAX_ROWS = 1000
TEMPLATE_INPUT_MAX_ROW = IMPORT_MAX_ROWS + 1


def generate_template(
    region_names: Sequence[str] | None = None,
    plane_type_names: Sequence[str] | None = None,
) -> io.BytesIO:
    """生成 Excel 导入模板。

    Args:
        region_names: Region 名称下拉候选项。
        plane_type_names: 网络平面类型名称下拉候选项。

    Returns:
        BytesIO 对象，包含格式化后的 .xlsx 模板文件。
    """
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "导入模板"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_idx - 1]

    ws.freeze_panes = "A2"
    _add_template_dropdowns(wb, ws, region_names, plane_type_names)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _add_template_dropdowns(
    wb: Workbook,
    ws: Any,
    region_names: Sequence[str] | None,
    plane_type_names: Sequence[str] | None,
) -> None:
    """给模板字段添加基于隐藏候选项工作表的数据验证下拉。"""
    option_groups = [
        ("RegionNameOptions", "A", _normalize_dropdown_options(region_names)),
        ("PlaneTypeNameOptions", "B", _normalize_dropdown_options(plane_type_names)),
    ]
    if not any(options for _, _, options in option_groups):
        return

    options_ws = wb.create_sheet(TEMPLATE_OPTION_SHEET)
    options_ws.sheet_state = "hidden"
    for _, column_letter, options in option_groups:
        for row_number, option in enumerate(options, 1):
            options_ws[f"{column_letter}{row_number}"] = option

    _add_dropdown_validation(wb, ws, "A", "RegionNameOptions", option_groups[0][2], "请选择已有 Region 名称")
    _add_dropdown_validation(wb, ws, "B", "PlaneTypeNameOptions", option_groups[1][2], "请选择已有网络平面类型")


def _normalize_dropdown_options(options: Sequence[str] | None) -> list[str]:
    """清理下拉候选项，保留首次出现顺序并去重。"""
    normalized = []
    seen = set()
    for option in options or []:
        value = str(option).strip()
        if not value or value in seen:
            continue
        normalized.append(value)
        seen.add(value)
    return normalized


def _add_dropdown_validation(
    wb: Workbook,
    ws: Any,
    column_letter: str,
    defined_name: str,
    options: list[str],
    error_message: str,
) -> None:
    if not options:
        return
    ref = f"'{TEMPLATE_OPTION_SHEET}'!${column_letter}$1:${column_letter}${len(options)}"
    wb.defined_names.add(DefinedName(defined_name, attr_text=ref))
    validation = DataValidation(
        type="list",
        formula1=f"={defined_name}",
        allow_blank=False,
        errorTitle="无效选项",
        error=error_message,
    )
    validation.errorStyle = "stop"
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{TEMPLATE_INPUT_MAX_ROW}")


def parse_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse an Excel file into a list of row dicts.

    Args:
        file_bytes: Excel 文件的二进制内容。

    Returns:
        解析后的行数据列表，每行包含 region_name、plane_type_name、
        scope、ip_range、vlan_id、gateway_position、gateway_ip 等字段。
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ValueError("Excel 文件无法解析，请确认文件为有效的 .xlsx 工作簿") from exc
    ws = cast(Worksheet, wb.active)
    rows = []
    header_values = [str(cell.value or "").strip() for cell in ws[1]]
    if header_values[: len(TEMPLATE_HEADERS)] != TEMPLATE_HEADERS:
        wb.close()
        raise ValueError("Excel 模板表头不匹配，请下载最新导入模板后重试")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        scope = str(row[2] or "").strip() or None
        ip_range_index = 3
        vlan_id_index = 4
        gateway_position_index = 5
        gateway_ip_index = 6
        vlan_id = row[vlan_id_index] if len(row) > vlan_id_index else None
        if isinstance(vlan_id, str):
            vlan_id = vlan_id.strip() or None
        rows.append(
            {
                "row_number": row_idx,
                "region_name": str(row[0] or "").strip(),
                "plane_type_name": str(row[1] or "").strip(),
                "scope": scope,
                "ip_range": str(row[ip_range_index] or "").strip(),
                "vlan_id": vlan_id,
                "gateway_position": (
                    str(row[gateway_position_index] or "").strip() if len(row) > gateway_position_index else ""
                )
                or None,
                "gateway_ip": (str(row[gateway_ip_index] or "").strip() if len(row) > gateway_ip_index else "") or None,
            }
        )
    wb.close()
    return rows


def build_export(data: list[dict[str, Any]]) -> io.BytesIO:
    """Build an Excel export workbook from Region network plane data.

    Args:
        data: 导出数据列表，每行应包含 region_name、plane_type_name、
              parent_plane_type_name、scope、is_private、vrf、ip_range、
              vlan_id、gateway_position、gateway_ip、updated_at。

    Returns:
        BytesIO 对象，包含格式化后的 .xlsx 导出文件。
    """
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "网络平面导出"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = EXPORT_COLUMN_WIDTHS[col_idx - 1]

    for row_idx, item in enumerate(data, 2):
        values = [
            item.get("region_name", ""),
            item.get("plane_type_name", ""),
            item.get("parent_plane_type_name", ""),
            item.get("scope", "Global"),
            item.get("is_private", ""),
            item.get("vrf", ""),
            item.get("ip_range", ""),
            item.get("vlan_id"),
            item.get("gateway_position", ""),
            item.get("gateway_ip", ""),
            item.get("updated_at", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

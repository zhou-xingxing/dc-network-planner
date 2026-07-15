"""Excel 导入导出接口测试。"""

import io
from typing import cast

from cachetools import TTLCache
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.user import User
from app.services import excel as excel_service
from app.utils.excel_utils import IMPORT_MAX_ROWS, TEMPLATE_HEADERS


def _workbook_bytes(headers, rows):
    """生成测试导入所需的 xlsx 二进制内容。"""
    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    return buffer.getvalue()


IMPORT_TEMPLATE_HEADERS = list(TEMPLATE_HEADERS)


def _setup_import_target(client, admin_headers, user_headers_factory):
    """创建导入确认所需的 Region、PlaneType 和授权用户。"""
    region = client.post("/api/regions", json={"name": "导入区域"}, headers=admin_headers).json()
    plane_type = client.post("/api/network-plane-types", json={"name": "导入平面"}, headers=admin_headers).json()
    user_headers = user_headers_factory([region["id"]])
    return region, plane_type, user_headers


def test_import_preview_region_ids_do_not_consume_cache(monkeypatch):
    """确认导入前的 Region 权限预检查只查看缓存，不消费预览数据。"""
    monkeypatch.setattr(excel_service, "_import_cache", TTLCache(maxsize=10, ttl=60))
    rows = [{"_region_id": "region-a"}, {"_region_id": "region-b"}]

    preview_id = excel_service.store_preview(rows)

    assert excel_service.get_preview_region_ids(preview_id) == {"region-a", "region-b"}
    assert excel_service.get_preview(preview_id) == rows
    assert excel_service.consume_preview(preview_id) == rows
    assert excel_service.get_preview(preview_id) is None


def test_import_preview_cache_expires(monkeypatch):
    """导入预览缓存超过 TTL 后应不可读取，并可主动清理。"""
    monkeypatch.setattr(excel_service, "_import_cache", TTLCache(maxsize=10, ttl=0))
    preview_id = excel_service.store_preview([{"_region_id": "region-a"}])

    assert excel_service.cleanup_expired_previews() == 1
    assert excel_service.get_preview(preview_id) is None


def test_import_preview_cache_respects_maxsize(monkeypatch):
    """导入预览缓存超过容量上限后应淘汰较早的预览。"""
    monkeypatch.setattr(excel_service, "_import_cache", TTLCache(maxsize=1, ttl=60))

    first_preview_id = excel_service.store_preview([{"_region_id": "region-a"}])
    second_rows = [{"_region_id": "region-b"}]
    second_preview_id = excel_service.store_preview(second_rows)

    assert excel_service.get_preview(first_preview_id) is None
    assert excel_service.get_preview(second_preview_id) == second_rows


def test_administrator_cannot_use_excel_import(client, admin_headers):
    """管理员不可使用 Excel 导入预览和确认接口。"""
    preview_response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                _workbook_bytes(IMPORT_TEMPLATE_HEADERS, []),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=admin_headers,
    )
    confirm_response = client.post(
        "/api/excel/import/confirm",
        json={"preview_id": "not-used"},
        headers=admin_headers,
    )

    assert preview_response.status_code == 403
    assert preview_response.json()["detail"] == "administrator 不可使用 Excel 导入功能"
    assert confirm_response.status_code == 403
    assert confirm_response.json()["detail"] == "administrator 不可使用 Excel 导入功能"


def test_preview_import_rejects_xls_extension(client, user_headers_factory):
    """导入预览只允许上传 .xlsx 文件。"""
    user_headers = user_headers_factory([])
    response = client.post(
        "/api/excel/import/preview",
        files={"file": ("import.xls", b"not-a-real-xls", "application/vnd.ms-excel")},
        headers=user_headers,
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "仅支持 .xlsx 文件"


def test_preview_import_rejects_invalid_xlsx(client, user_headers_factory):
    """损坏的 xlsx 应返回可理解的 400 错误。"""
    user_headers = user_headers_factory([])
    response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                b"not-a-workbook",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Excel 文件无法解析，请确认文件为有效的 .xlsx 工作簿"


def test_preview_import_rejects_rows_over_limit(client, user_headers_factory):
    """导入数据超过 1000 条时应拒绝整份文件。"""
    user_headers = user_headers_factory([])
    rows = [
        ["导入区域", "导入平面", "Global", "10.10.0.0/24", 100, "CE01", "10.10.0.1"] for _ in range(IMPORT_MAX_ROWS + 1)
    ]

    response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                _workbook_bytes(IMPORT_TEMPLATE_HEADERS, rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "单次最多导入 1000 条数据，当前文件包含 1001 条"


def test_preview_import_accepts_rows_at_limit(client, user_headers_factory):
    """导入数据恰好 1000 条时应继续执行逐行预览校验。"""
    user_headers = user_headers_factory([])
    rows = [
        ["不存在区域", "不存在平面", "Global", "10.10.0.0/24", 100, "CE01", "10.10.0.1"] for _ in range(IMPORT_MAX_ROWS)
    ]

    response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                _workbook_bytes(IMPORT_TEMPLATE_HEADERS, rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )

    assert response.status_code == 200
    assert response.json()["total_rows"] == IMPORT_MAX_ROWS
    assert response.json()["valid_rows"] == 0
    assert len(response.json()["error_rows"]) == IMPORT_MAX_ROWS


def test_preview_import_rejects_gateway_ip_outside_cidr(client, admin_headers, user_headers_factory):
    """网关 IP 不在 CIDR 范围内时，预览阶段应提前标记为校验错误。"""
    _, _, user_headers = _setup_import_target(client, admin_headers, user_headers_factory)
    file_bytes = _workbook_bytes(
        IMPORT_TEMPLATE_HEADERS,
        [["导入区域", "导入平面", "Global", "10.10.0.0/24", 100, "CE01", "10.10.1.1"]],
    )

    response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["valid_rows"] == 0
    assert data["error_rows"][0]["error_type"] == "validation"
    assert "网关 IP 10.10.1.1 必须在平面 CIDR 10.10.0.0/24 范围内" in data["error_rows"][0]["errors"]


def test_preview_import_rows_only_include_valid_rows(client, admin_headers, user_headers_factory):
    """预览响应 rows 只返回确认导入会使用的有效行。"""
    _, _, user_headers = _setup_import_target(client, admin_headers, user_headers_factory)
    file_bytes = _workbook_bytes(
        IMPORT_TEMPLATE_HEADERS,
        [
            ["导入区域", "导入平面", "", "10.10.0.0/24", 100, "CE01", "10.10.0.1"],
            ["导入区域", "导入平面", "Global", "bad-cidr", 101, "CE02", "10.10.1.1"],
        ],
    )

    response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["total_rows"] == 2
    assert data["valid_rows"] == 1
    assert [row["row_number"] for row in data["rows"]] == [2]
    assert data["rows"][0]["scope"] == "Global"
    assert data["error_rows"][0]["row"] == 3
    assert data["error_rows"][0]["region_name"] == "导入区域"
    assert data["error_rows"][0]["error_type"] == "validation"


def test_preview_import_marks_all_vlan_value_errors_as_validation(client, admin_headers, user_headers_factory):
    """VLAN 小数、字符串和越界值都应作为校验错误返回。"""
    _, _, user_headers = _setup_import_target(client, admin_headers, user_headers_factory)
    file_bytes = _workbook_bytes(
        IMPORT_TEMPLATE_HEADERS,
        [
            ["导入区域", "导入平面", "Global", "10.10.0.0/24", 10.5, "CE01", "10.10.0.1"],
            ["导入区域", "导入平面", "业务可用区一", "10.10.1.0/24", "abc", "CE02", "10.10.1.1"],
            ["导入区域", "导入平面", "业务可用区二", "10.10.2.0/24", 4095, "CE03", "10.10.2.1"],
        ],
    )

    response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["valid_rows"] == 0
    assert data["rows"] == []
    assert [row["row"] for row in data["error_rows"]] == [2, 3, 4]
    assert [row["region_name"] for row in data["error_rows"]] == ["导入区域", "导入区域", "导入区域"]
    assert [row["error_type"] for row in data["error_rows"]] == ["validation", "validation", "validation"]
    errors = [row["errors"][0] for row in data["error_rows"]]
    assert errors == ["无效 VLAN ID: 10.5", "无效 VLAN ID: abc", "无效 VLAN ID: 4095"]


def test_preview_import_marks_unpermitted_region_as_not_importable(client, admin_headers, user_headers_factory):
    """预览阶段标注未授权 Region 行，不放入可确认导入的有效行。"""
    region = client.post("/api/regions", json={"name": "导入区域"}, headers=admin_headers).json()
    client.post("/api/regions", json={"name": "其他区域"}, headers=admin_headers).json()
    client.post("/api/network-plane-types", json={"name": "导入平面"}, headers=admin_headers).json()
    user_headers = user_headers_factory([region["id"]])
    file_bytes = _workbook_bytes(
        IMPORT_TEMPLATE_HEADERS,
        [
            ["导入区域", "导入平面", "Global", "10.10.0.0/24", 100, "CE01", "10.10.0.1"],
            ["其他区域", "导入平面", "Global", "10.20.0.0/24", 101, "CE02", "10.20.0.1"],
        ],
    )

    response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )

    assert response.status_code == 200
    data = response.json()
    assert data["total_rows"] == 2
    assert data["valid_rows"] == 1
    assert [row["row_number"] for row in data["rows"]] == [2]
    assert data["error_rows"][0]["row"] == 3
    assert data["error_rows"][0]["region_name"] == "其他区域"
    assert data["error_rows"][0]["error_type"] == "permission"
    assert "用户未授权管理此 Region：其他区域，仅提供预览功能，不能实际导入" in data["error_rows"][0]["errors"]


def test_confirm_import_rechecks_region_permission(test_db, client, admin_headers, user_headers_factory):
    """确认导入阶段应继续按当前权限二次校验 Region 写权限。"""
    region, _, user_headers = _setup_import_target(client, admin_headers, user_headers_factory)
    file_bytes = _workbook_bytes(
        IMPORT_TEMPLATE_HEADERS,
        [["导入区域", "导入平面", "Global", "10.10.0.0/24", 100, "CE01", "10.10.0.1"]],
    )
    preview_response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )
    assert preview_response.status_code == 200
    assert preview_response.json()["valid_rows"] == 1

    session = Session(test_db)
    try:
        user = session.query(User).filter(User.username == "region-user").one()
        user.region_permissions.clear()
        session.commit()
    finally:
        session.close()

    response = client.post(
        "/api/excel/import/confirm",
        json={"preview_id": preview_response.json()["preview_id"]},
        headers=user_headers,
    )

    assert region["name"] == "导入区域"
    assert response.status_code == 403
    assert response.json()["detail"] == "无权管理该 Region 的业务数据"


def test_confirm_import_success_creates_region_plane(client, admin_headers, user_headers_factory):
    """确认导入成功后应创建 Region 网络平面。"""
    region, _, user_headers = _setup_import_target(client, admin_headers, user_headers_factory)
    file_bytes = _workbook_bytes(
        IMPORT_TEMPLATE_HEADERS,
        [["导入区域", "导入平面", "", "10.10.0.0/24", 100, "CE01", "10.10.0.1"]],
    )
    preview_response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )
    assert preview_response.status_code == 200
    preview_data = preview_response.json()
    assert preview_data["valid_rows"] == 1

    response = client.post(
        "/api/excel/import/confirm",
        json={"preview_id": preview_data["preview_id"]},
        headers=user_headers,
    )

    assert response.status_code == 200
    result = response.json()
    assert result["success"] is True
    assert result["imported_count"] == 1
    assert result["error_count"] == 0
    assert result["errors"] == []
    assert len(result["row_results"]) == 1
    assert result["row_results"][0] == {
        "row": 2,
        "status": "success",
        "region_name": "导入区域",
        "plane_type_name": "导入平面",
        "scope": "Global",
        "ip_range": "10.10.0.0/24",
        "vlan_id": 100,
        "gateway_position": "CE01",
        "gateway_ip": "10.10.0.1",
        "plane_id": result["row_results"][0]["plane_id"],
        "errors": [],
    }
    assert result["row_results"][0]["plane_id"]

    second_response = client.post(
        "/api/excel/import/confirm",
        json={"preview_id": preview_data["preview_id"]},
        headers=user_headers,
    )
    assert second_response.status_code == 200
    assert second_response.json()["success"] is False
    assert second_response.json()["imported_count"] == 0
    assert second_response.json()["errors"][0]["errors"] == ["预览数据已过期，请重新上传"]
    assert second_response.json()["row_results"] == []

    planes_response = client.get(f"/api/regions/{region['id']}/planes", headers=user_headers)
    assert planes_response.status_code == 200
    planes = planes_response.json()
    assert len(planes) == 1
    assert planes[0]["plane_type_name"] == "导入平面"
    assert planes[0]["scope"] == "Global"
    assert planes[0]["cidr"] == "10.10.0.0/24"
    assert planes[0]["vlan_id"] == 100
    assert planes[0]["gateway_position"] == "CE01"
    assert planes[0]["gateway_ip"] == "10.10.0.1"


def test_confirm_import_returns_success_and_failure_details(client, admin_headers, user_headers_factory):
    """确认导入应返回每个有效行的最终成功或失败结果。"""
    _, _, user_headers = _setup_import_target(client, admin_headers, user_headers_factory)
    file_bytes = _workbook_bytes(
        IMPORT_TEMPLATE_HEADERS,
        [
            ["导入区域", "导入平面", "Global", "10.10.0.0/24", 100, "CE01", "10.10.0.1"],
            ["导入区域", "导入平面", "Global", "10.10.1.0/24", 101, "CE02", "10.10.1.1"],
        ],
    )
    preview_response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )
    assert preview_response.status_code == 200
    assert preview_response.json()["valid_rows"] == 2

    response = client.post(
        "/api/excel/import/confirm",
        json={"preview_id": preview_response.json()["preview_id"]},
        headers=user_headers,
    )

    assert response.status_code == 200
    result = response.json()
    assert result["success"] is False
    assert result["imported_count"] == 1
    assert result["error_count"] == 1
    assert [row["status"] for row in result["row_results"]] == ["success", "failed"]
    assert [row["row"] for row in result["row_results"]] == [2, 3]
    assert result["row_results"][0]["plane_id"]
    assert result["row_results"][0]["errors"] == []
    assert result["row_results"][1]["plane_id"] is None
    assert result["row_results"][1]["errors"] == ["该网络平面类型已在本Region 的 Global 作用域中创建，不能重复创建"]
    assert result["errors"][0]["row"] == 3
    assert result["errors"][0]["errors"] == result["row_results"][1]["errors"]


def test_confirm_import_without_valid_rows_is_not_successful(client, admin_headers, user_headers_factory):
    """预览中没有有效行时，直接调用确认接口也不应返回成功。"""
    _, _, user_headers = _setup_import_target(client, admin_headers, user_headers_factory)
    file_bytes = _workbook_bytes(
        IMPORT_TEMPLATE_HEADERS,
        [["导入区域", "导入平面", "Global", "bad-cidr", 100, "CE01", "10.10.0.1"]],
    )
    preview_response = client.post(
        "/api/excel/import/preview",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=user_headers,
    )
    assert preview_response.status_code == 200
    assert preview_response.json()["valid_rows"] == 0

    response = client.post(
        "/api/excel/import/confirm",
        json={"preview_id": preview_response.json()["preview_id"]},
        headers=user_headers,
    )

    assert response.status_code == 200
    assert response.json() == {
        "success": False,
        "imported_count": 0,
        "error_count": 0,
        "errors": [],
        "row_results": [],
    }


def test_download_template_includes_region_and_plane_type_dropdowns(client, admin_headers):
    """导入模板应包含当前全部 Region 和网络平面类型作为下拉候选项。"""
    client.post("/api/regions", json={"name": "区域二"}, headers=admin_headers)
    client.post("/api/regions", json={"name": "区域一"}, headers=admin_headers)
    client.post("/api/network-plane-types", json={"name": "业务平面"}, headers=admin_headers)
    client.post("/api/network-plane-types", json={"name": "管理平面"}, headers=admin_headers)

    response = client.get("/api/excel/template", headers=admin_headers)

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    worksheet = cast(Worksheet, workbook.active)
    options = workbook["候选项"]
    assert options.sheet_state == "hidden"
    assert [options.cell(row=row_number, column=1).value for row_number in range(1, 3)] == ["区域一", "区域二"]
    assert [options.cell(row=row_number, column=2).value for row_number in range(1, 3)] == ["业务平面", "管理平面"]
    assert worksheet.title == "导入模板"
    workbook.close()


def test_export_orders_region_planes_by_names(client, admin_headers, user_headers_factory):
    """Excel 导出默认按 Region 名称、网络平面名称、scope 和 CIDR 升序。"""
    region_b = client.post("/api/regions", json={"name": "区域二"}, headers=admin_headers).json()
    region_a = client.post("/api/regions", json={"name": "区域一"}, headers=admin_headers).json()
    plane_type = client.post("/api/network-plane-types", json={"name": "管理平面"}, headers=admin_headers).json()
    user_headers = user_headers_factory([region_a["id"], region_b["id"]])
    client.post(
        f"/api/regions/{region_b['id']}/planes",
        json={"plane_type_id": plane_type["id"], "cidr": "10.0.2.0/24"},
        headers=user_headers,
    )
    client.post(
        f"/api/regions/{region_a['id']}/planes",
        json={"plane_type_id": plane_type["id"], "cidr": "10.0.1.0/24"},
        headers=user_headers,
    )

    response = client.get("/api/excel/export", headers=admin_headers)

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    worksheet = cast(Worksheet, workbook.active)
    region_names = [worksheet.cell(row=row_number, column=1).value for row_number in range(2, 4)]
    workbook.close()
    assert region_names == ["区域一", "区域二"]


def test_export_filters_region_planes_by_region(client, admin_headers, user_headers_factory):
    """Excel 导出支持按 Region 筛选网络平面。"""
    region_a = client.post("/api/regions", json={"name": "区域一"}, headers=admin_headers).json()
    region_b = client.post("/api/regions", json={"name": "区域二"}, headers=admin_headers).json()
    plane_type = client.post("/api/network-plane-types", json={"name": "管理平面"}, headers=admin_headers).json()
    user_headers = user_headers_factory([region_a["id"], region_b["id"]])
    client.post(
        f"/api/regions/{region_a['id']}/planes",
        json={"plane_type_id": plane_type["id"], "cidr": "10.0.1.0/24"},
        headers=user_headers,
    )
    client.post(
        f"/api/regions/{region_b['id']}/planes",
        json={"plane_type_id": plane_type["id"], "cidr": "10.0.2.0/24"},
        headers=user_headers,
    )

    response = client.get(f"/api/excel/export?region_id={region_a['id']}", headers=admin_headers)

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    worksheet = cast(Worksheet, workbook.active)
    region_names = [worksheet.cell(row=row_number, column=1).value for row_number in range(2, worksheet.max_row + 1)]
    workbook.close()
    assert region_names == ["区域一"]


def test_export_includes_all_region_plane_fields(client, admin_headers, user_headers_factory):
    """Excel 导出应包含 Region 网络平面的全部字段。"""
    region = client.post("/api/regions", json={"name": "区域一"}, headers=admin_headers).json()
    parent_type = client.post(
        "/api/network-plane-types",
        json={"name": "父平面", "is_private": True, "vrf": "VRF-PARENT"},
        headers=admin_headers,
    ).json()
    child_type = client.post(
        "/api/network-plane-types",
        json={"name": "子平面", "parent_id": parent_type["id"], "is_private": True, "vrf": "VRF-CHILD"},
        headers=admin_headers,
    ).json()
    user_headers = user_headers_factory([region["id"]])
    client.post(
        f"/api/regions/{region['id']}/planes",
        json={"plane_type_id": parent_type["id"], "cidr": "10.0.0.0/16"},
        headers=user_headers,
    )
    client.post(
        f"/api/regions/{region['id']}/planes",
        json={
            "plane_type_id": child_type["id"],
            "cidr": "10.0.1.0/24",
            "vlan_id": 200,
            "gateway_position": "SW01 / SW02",
            "gateway_ip": "10.0.1.1",
        },
        headers=user_headers,
    )

    response = client.get("/api/excel/export", headers=admin_headers)

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    worksheet = cast(Worksheet, workbook.active)
    headers = [cell.value for cell in worksheet[1]]
    child_row = next(row for row in worksheet.iter_rows(min_row=2, values_only=True) if row[1] == "子平面")
    workbook.close()
    assert headers == [
        "区域",
        "网络平面类型",
        "父级网络平面类型",
        "作用域",
        "是否私网",
        "VRF",
        "IP地址段(CIDR)",
        "VLAN ID",
        "网关位置",
        "网关IP",
        "更新时间",
    ]
    assert child_row[0] == "区域一"
    assert child_row[1] == "子平面"
    assert child_row[2] == "父平面"
    assert child_row[3] == "Global"
    assert child_row[4] == "是"
    assert child_row[5] == "VRF-CHILD"
    assert child_row[6] == "10.0.1.0/24"
    assert child_row[7] == 200
    assert child_row[8] == "SW01 / SW02"
    assert child_row[9] == "10.0.1.1"
    assert child_row[10]

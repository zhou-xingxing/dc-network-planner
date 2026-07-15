"""Excel 工具函数测试。"""

import io
from typing import cast

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.excel_utils import IMPORT_MAX_ROWS, TEMPLATE_INPUT_MAX_ROW, generate_template, parse_excel


def test_generate_template():
    """生成的 Excel 导入模板应包含预期工作表和表头。"""
    buf = generate_template(region_names=["区域一", "区域二"], plane_type_names=["管理平面", "业务平面"])
    wb = load_workbook(buf)
    ws = cast(Worksheet, wb.active)

    assert ws.title == "导入模板"
    headers = [cell.value for cell in ws[1]]
    assert "区域名称\n(必填)" in headers
    assert "网络平面类型\n(必填)" in headers
    assert "作用域\n(选填，空值默认 Global)" in headers
    assert "IP地址段(CIDR)\n(必填，如 10.0.0.0/24)" in headers
    assert "VLAN ID\n(选填，1-4094)" in headers
    assert "网关位置\n(选填)" in headers
    assert "网关IP\n(选填)" in headers
    assert "候选项" in wb.sheetnames
    assert wb["候选项"].sheet_state == "hidden"
    assert wb["候选项"]["A1"].value == "区域一"
    assert wb["候选项"]["B1"].value == "管理平面"
    validations = list(ws.data_validations.dataValidation)
    assert len(validations) == 2
    assert {validation.formula1 for validation in validations} == {"=RegionNameOptions", "=PlaneTypeNameOptions"}
    assert TEMPLATE_INPUT_MAX_ROW == IMPORT_MAX_ROWS + 1
    assert {str(validation.sqref) for validation in validations} == {"A2:A1001", "B2:B1001"}

    wb.close()


def test_generate_template_without_dropdown_options():
    """没有候选项时仍应正常生成基础导入模板。"""
    buf = generate_template(region_names=[], plane_type_names=[])
    wb = load_workbook(buf)
    ws = cast(Worksheet, wb.active)

    assert ws.title == "导入模板"
    assert "候选项" not in wb.sheetnames
    assert list(ws.data_validations.dataValidation) == []

    wb.close()


def test_generate_template_deduplicates_and_filters_dropdown_options():
    """下拉候选项应过滤空值并按首次出现顺序去重。"""
    buf = generate_template(
        region_names=["区域一", "", " 区域一 ", "区域二"],
        plane_type_names=["管理平面", " ", "管理平面", "业务平面"],
    )
    wb = load_workbook(buf)
    options = wb["候选项"]

    assert [options.cell(row=row_number, column=1).value for row_number in range(1, 3)] == ["区域一", "区域二"]
    assert [options.cell(row=row_number, column=2).value for row_number in range(1, 3)] == ["管理平面", "业务平面"]

    wb.close()


def test_parse_excel_rejects_mismatched_template_headers():
    """导入解析只接受当前模板表头。"""
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.append(["区域名称", "网络平面类型", "作用域", "CIDR", "VLAN ID", "网关位置", "网关IP"])
    ws.append(["测试区域", "管理平面", "10.0.0.0/24", 100, "SW01 / SW02", "10.0.0.1"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    with pytest.raises(ValueError, match="Excel 模板表头不匹配"):
        parse_excel(output.getvalue())
    wb.close()


def test_parse_empty_excel():
    """没有数据行的模板应解析为空列表。"""
    buf = generate_template()
    rows = parse_excel(buf.getvalue())
    assert rows == []


def test_parse_excel_reads_rows_from_generated_template():
    """导入解析应读取当前模板中的用户填写行。"""
    buf = generate_template()
    wb = load_workbook(buf)
    ws = cast(Worksheet, wb.active)
    ws.append(["测试区域", "管理平面", "", "10.0.0.0/24", 100, "SW01 / SW02", "10.0.0.1"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    rows = parse_excel(output.getvalue())
    assert len(rows) == 1
    assert rows[0]["region_name"] == "测试区域"
    assert rows[0]["scope"] is None
    assert rows[0]["ip_range"] == "10.0.0.0/24"
    assert rows[0]["vlan_id"] == 100
    assert rows[0]["gateway_position"] == "SW01 / SW02"
    assert rows[0]["gateway_ip"] == "10.0.0.1"
    wb.close()


def test_parse_excel_trims_cell_values_and_normalizes_blank_cells():
    """导入解析应清理单元格左右空格，并将空白单元格转换为 None。"""
    buf = generate_template()
    wb = load_workbook(buf)
    ws = cast(Worksheet, wb.active)
    ws.append([" 测试区域 ", " 管理平面 ", " 业务可用区一 ", " 10.0.0.0/24 ", " 100 ", " SW01 / SW02 ", " 10.0.0.1 "])
    ws.append(["测试区域", "管理平面", "   ", "10.0.1.0/24", "   ", "   ", "   "])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    rows = parse_excel(output.getvalue())
    assert rows[0]["region_name"] == "测试区域"
    assert rows[0]["plane_type_name"] == "管理平面"
    assert rows[0]["scope"] == "业务可用区一"
    assert rows[0]["ip_range"] == "10.0.0.0/24"
    assert rows[0]["vlan_id"] == "100"
    assert rows[0]["gateway_position"] == "SW01 / SW02"
    assert rows[0]["gateway_ip"] == "10.0.0.1"
    assert rows[1]["scope"] is None
    assert rows[1]["vlan_id"] is None
    assert rows[1]["gateway_position"] is None
    assert rows[1]["gateway_ip"] is None
    wb.close()
